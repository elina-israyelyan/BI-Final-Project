import os

import numpy as np
import pandas as pd
import pyodbc
from openpyxl import load_workbook

import readconfig

# Set the working directory to the location of scripts
path = r"/home/elina/Desktop/ayodas/bi_group_project/"
os.chdir(path)


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def populate_ER(db='Orders_RELATIONAL_DB', src='data/raw_data_source.xlsx'):
    """
    Populate The ER table from the source file
    """
    # Call to read the configuration file
    c_ER = readconfig.get_sql_config(r'sql_server_config.cfg', "Database1")
    # Create a connection string for SQL Server
    conn_info_ER = 'Driver={};Server={};Database={};Trusted_Connection={};UID={};PWD={};'.format(*c_ER)
    # Connect to the server and to the desired database
    conn_ER = pyodbc.connect(conn_info_ER)
    # Create a Cursor class instance for executing T-SQL statements
    cursor_ER = conn_ER.cursor()

    cursor_ER.execute(f'use {db}')
    table_names = get_sheetnames_xlsx(src)
    for sheet in table_names:
        cursor_ER.execute(f'ALTER TABLE {sheet} NOCHECK CONSTRAINT ALL;')
        sheet_data = pd.read_excel(src, sheet_name=sheet)
        sheet_data = sheet_data.replace({np.nan: None})

        for _, row in sheet_data.iterrows():
            row_as_dict = row.to_dict()
            placeholders = ', '.join(['?'] * len(row_as_dict))
            columns = ', '.join(row_as_dict.keys())
            sql_query = """INSERT INTO %s ( %s ) VALUES ( %s )""" % (sheet, columns, placeholders)
            print(row_as_dict)
            cursor_ER.execute(sql_query, list(row_as_dict.values()))

    for sheet in table_names:
        cursor_ER.execute(f"ALTER TABLE {sheet} WITH CHECK CHECK CONSTRAINT ALL;")

    cursor_ER.commit()
    cursor_ER.close()
    conn_ER.close()


if __name__ == '__main__':
    populate_ER(db='Orders_RELATIONAL_DB', src='data/raw_data_source.xlsx')
